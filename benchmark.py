"""
Benchmark runner — executes each algorithm on the same scenario
and collects performance metrics.
"""

from __future__ import annotations

import time
import numpy as np
from dataclasses import dataclass, field
from typing import Dict, List

import config as cfg
from environment import Environment
from algorithms.base import BaseAlgorithm


@dataclass
class RunMetrics:
    algorithm: str
    complexity: str
    total_collisions: int = 0
    collision_rate: float = 0.0       # collisions / agent
    agents_arrived: int = 0
    success_rate: float = 0.0         # arrived / total
    makespan: float = 0.0             # time until last arrival (or timeout)
    avg_arrival_time: float = 0.0
    avg_speed_utilization: float = 0.0  # mean(actual_speed / preferred) over all steps
    computation_time: float = 0.0     # wall-clock seconds for algorithm calls
    total_sim_time: float = 0.0       # wall-clock seconds for entire run
    steps: int = 0


def run_benchmark(
    env: Environment,
    algo: BaseAlgorithm,
    verbose: bool = False,
) -> RunMetrics:
    """Run a single algorithm on the environment and return metrics."""
    env.reset()
    n = env.n

    metrics = RunMetrics(algorithm=algo.name, complexity=algo.complexity)
    speed_utils: List[float] = []

    t_start = time.perf_counter()
    algo_time = 0.0
    step = 0

    while not env.all_arrived() and not env.timed_out():
        positions = env.get_positions()
        headings = env.get_headings()
        speeds = env.get_speeds()
        preferred = env.get_preferred_speeds()
        active = env.get_active_mask()

        # time the algorithm
        t0 = time.perf_counter()
        commanded = algo.compute_speeds(positions, headings, speeds, preferred, active, cfg.DT)
        t1 = time.perf_counter()
        algo_time += (t1 - t0)

        env.step(commanded)
        step += 1

        # speed utilization for active agents
        active_mask = env.get_active_mask()
        cur_speeds = env.get_speeds()
        if np.any(active_mask):
            util = np.mean(cur_speeds[active_mask] / cfg.PREFERRED_SPEED)
            speed_utils.append(util)

        if verbose and step % 500 == 0:
            n_active = int(np.sum(active_mask))
            print(f"  [{algo.name}] t={env.t:.1f}s  active={n_active}  collisions={env.collisions}")

    t_end = time.perf_counter()

    # gather results
    metrics.steps = step
    metrics.total_collisions = env.collisions
    metrics.collision_rate = env.collisions / max(n, 1)
    metrics.agents_arrived = int(np.sum(env.arrived))
    metrics.success_rate = metrics.agents_arrived / max(n, 1)
    metrics.total_sim_time = t_end - t_start
    metrics.computation_time = algo_time

    arrived_mask = env.arrived & ~np.isnan(env.arrival_times)
    if np.any(arrived_mask):
        metrics.makespan = float(np.nanmax(env.arrival_times[arrived_mask]))
        metrics.avg_arrival_time = float(np.nanmean(env.arrival_times[arrived_mask]))
    else:
        metrics.makespan = cfg.MAX_TIME

    metrics.avg_speed_utilization = float(np.mean(speed_utils)) if speed_utils else 0.0

    return metrics


def format_results_table(results: List[RunMetrics]) -> str:
    """Format benchmark results as a readable table."""
    from tabulate import tabulate

    headers = [
        "Algorithm",
        "Complexity",
        "Collisions",
        "Coll. Rate",
        "Success %",
        "Makespan (s)",
        "Avg Arrival (s)",
        "Speed Util.",
        "Algo Time (s)",
        "Total Time (s)",
    ]

    rows = []
    for r in results:
        rows.append([
            r.algorithm,
            r.complexity,
            r.total_collisions,
            f"{r.collision_rate:.4f}",
            f"{r.success_rate * 100:.1f}%",
            f"{r.makespan:.1f}",
            f"{r.avg_arrival_time:.1f}",
            f"{r.avg_speed_utilization:.3f}",
            f"{r.computation_time:.3f}",
            f"{r.total_sim_time:.1f}",
        ])

    return tabulate(rows, headers=headers, tablefmt="grid")


def export_to_excel(results: List[RunMetrics], path: str):
    """Export results to an Excel file with formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark Results"

    headers = [
        "Алгоритм", "Асимптотическая сложность", "Всего коллизий",
        "Частота коллизий", "Успешность (%)", "Макс. время доставки (с)",
        "Среднее время прибытия (с)", "Утилизация скорости", "Время работы алгоритма (с)",
        "Общее время симуляции (с)", "Шагов",
    ]

    # header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # data rows
    data_align = Alignment(horizontal="center", vertical="center")
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    for row_idx, r in enumerate(results, 2):
        values = [
            r.algorithm,
            r.complexity,
            r.total_collisions,
            round(r.collision_rate, 4),
            round(r.success_rate * 100, 1),
            round(r.makespan, 1),
            round(r.avg_arrival_time, 1),
            round(r.avg_speed_utilization, 3),
            round(r.computation_time, 3),
            round(r.total_sim_time, 1),
            r.steps,
        ]
        fill = alt_fill if row_idx % 2 == 0 else None
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = data_align
            cell.border = thin_border
            if fill:
                cell.fill = fill

    # auto-width
    for col in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, len(results) + 2)
        )
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(max_len + 4, 14)

    # freeze header row
    ws.freeze_panes = "A2"

    # =====================================================================
    # Sheet 2: Параметры симуляции
    # =====================================================================
    ws2 = wb.create_sheet("Параметры симуляции")

    section_font = Font(bold=True, size=12, color="FFFFFF")
    section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    label_font = Font(bold=True, size=11)
    value_font = Font(size=11)
    wrap_align = Alignment(vertical="top", wrap_text=True)

    import config as cfg_mod

    params = [
        ("Параметр", "Значение"),
        ("Размер мира", f"{cfg_mod.WORLD_SIZE:.0f} x {cfg_mod.WORLD_SIZE:.0f} м ({cfg_mod.WORLD_SIZE/1000:.0f} x {cfg_mod.WORLD_SIZE/1000:.0f} км)"),
        ("Количество агентов", f"{cfg_mod.NUM_AGENTS}"),
        ("Шаг симуляции (dt)", f"{cfg_mod.DT} с"),
        ("Макс. время симуляции", f"{cfg_mod.MAX_TIME:.0f} с ({cfg_mod.MAX_TIME/60:.0f} мин)"),
        ("Номинальная скорость", f"{cfg_mod.PREFERRED_SPEED} м/с"),
        ("Макс. скорость", f"{cfg_mod.MAX_SPEED} м/с"),
        ("Мин. скорость", f"{cfg_mod.MIN_SPEED} м/с"),
        ("Радиус агента", f"{cfg_mod.AGENT_RADIUS} м"),
        ("Защитный зазор", f"{cfg_mod.SAFETY_MARGIN} м"),
        ("Дистанция коллизии", f"{cfg_mod.COLLISION_DIST} м (центр-центр)"),
        ("Ветер", "Включён" if cfg_mod.WIND_ENABLED else "Выключен"),
        ("Базовая сила ветра", f"{cfg_mod.WIND_BASE_STRENGTH} м/с"),
        ("Макс. порыв ветра", f"{cfg_mod.WIND_GUST_STRENGTH} м/с"),
        ("Интервал смены ветра", f"{cfg_mod.WIND_CHANGE_INTERVAL} с"),
        ("Штраф уклонения", f"{cfg_mod.EVASION_SPEED_PENALTY * 100:.0f}% от скорости"),
        ("Random seed", f"{cfg_mod.SEED}"),
    ]

    for row_idx, (label, value) in enumerate(params, 1):
        c1 = ws2.cell(row=row_idx, column=1, value=label)
        c2 = ws2.cell(row=row_idx, column=2, value=value)
        c1.border = thin_border
        c2.border = thin_border
        if row_idx == 1:
            c1.font = section_font
            c1.fill = section_fill
            c2.font = section_font
            c2.fill = section_fill
            c1.alignment = header_align
            c2.alignment = header_align
        else:
            c1.font = label_font
            c2.font = value_font

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 45

    # =====================================================================
    # Sheet 3: Описание алгоритмов
    # =====================================================================
    ws3 = wb.create_sheet("Описание алгоритмов")

    algo_descriptions = [
        ("Алгоритм", "Тип", "Асимпт. сложность", "Принцип работы", "Особенности"),
        (
            "No-Control",
            "Baseline",
            "O(1)",
            "Каждый агент летит с номинальной скоростью без какого-либо управления.",
            "Базовый уровень для сравнения. Показывает количество коллизий при полном отсутствии деконфликтации.",
        ),
        (
            "FCFS",
            "Классический",
            "O(N log N)",
            "Приоритет по порядку (First-Come-First-Served). При сближении агент с меньшим приоритетом (больший ID) тормозит пропорционально расстоянию до опасной зоны.",
            "Простой, но не учитывает геометрию конфликта. Поиск соседей через KDTree.",
        ),
        (
            "VO (Velocity Obstacle)",
            "Классический",
            "O(N^2)",
            "Для каждого соседа строится конус опасных скоростей. Конус проецируется на 1D-линию (только скорость). Выбирается ближайшая к номинальной безопасная скорость вне заблокированных интервалов.",
            "Агент берёт на себя 100% ответственности за избегание. Может вызывать осцилляции при встречном движении.",
        ),
        (
            "VO-SSD (Speed Separation Detection)",
            "Классический",
            "O(N log N)",
            "Вариант VO, оптимизированный для speed-only контроля. Строит заблокированные интервалы скоростей на основе минимального расстояния сближения.",
            "Более точная параметризация конуса для скалярного управления.",
        ),
        (
            "RVO (Reciprocal VO)",
            "Классический",
            "O(N^2)",
            "Модификация VO с взаимным (reciprocal) разделением ответственности: каждый агент берёт на себя 50% корректировки.",
            "Устраняет осцилляции классического VO. Заблокированный интервал сдвигается к текущей скорости на 50%.",
        ),
        (
            "ORCA (Optimal Reciprocal Collision Avoidance)",
            "Классический",
            "O(N log N)",
            "Строит полуплоскости ORCA в 2D-пространстве скоростей. Каждая полуплоскость проецируется на линию heading-а, давая линейное ограничение s >= X или s <= X. Решается 1D линейная программа.",
            "Гарантирует безопасность при условии что все агенты используют ORCA. van den Berg et al. (2011).",
        ),
        (
            "CSORCA (Cooperative Speed ORCA)",
            "Классический",
            "O(N^2)",
            "Расширение ORCA с кооперативным распределением ответственности: агент с большим запасом скорости (далеко от min/max) берёт на себя большую долю корректировки.",
            "Лучший среди классических алгоритмов по количеству коллизий. Адаптивное разделение вместо фиксированных 50/50.",
        ),
        (
            "MILP (Mixed-Integer Linear Programming)",
            "Оптимизационный",
            "O(N^2 + LP)",
            "Формулирует назначение скоростей как задачу линейного программирования: минимизировать суммарное отклонение от номинала при ограничениях попарного разделения. Решается через scipy.optimize.linprog (HiGHS).",
            "Минимум коллизий, но слишком консервативный — агрессивно тормозит агентов, низкий success rate.",
        ),
        (
            "MAPPO (Multi-Agent PPO)",
            "MARL (нейросетевой)",
            "O(N*K*d)",
            "Shared actor-critic. Каждый агент наблюдает 10 ближайших соседей. 2-слойный MLP (tanh, hidden=64) выдаёт скорость. Централизованный критик при обучении.",
            "Необученная сеть (random weights). Базовый уровень для последующего обучения PPO. K=10 соседей, d=64.",
        ),
        (
            "MADDPG (Multi-Agent DDPG)",
            "MARL (нейросетевой)",
            "O(N*K*d^2)",
            "Decentralized actors + centralized critic. 3-слойный MLP (ReLU, hidden=128). Детерминистическая политика: obs -> speed.",
            "Необученная сеть. Базовый уровень для DDPG-обучения. Большая сеть (128 hidden) — дороже по compute.",
        ),
        (
            "MASAC (Multi-Agent SAC)",
            "MARL (нейросетевой)",
            "O(N*K*d^2)",
            "Стохастический актор: выдаёт mean + log_std гауссиана. При инференсе используется mean (детерминистически). Entropy-regularized.",
            "Необученная сеть. Базовый уровень для SAC-обучения. Та же архитектура что MADDPG + стохастическая голова.",
        ),
    ]

    for row_idx, row_data in enumerate(algo_descriptions, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_idx == 1:
                cell.font = section_font
                cell.fill = section_fill
                cell.alignment = header_align

    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 20
    ws3.column_dimensions["D"].width = 55
    ws3.column_dimensions["E"].width = 55

    # set row heights for readability
    for row_idx in range(2, len(algo_descriptions) + 1):
        ws3.row_dimensions[row_idx].height = 60

    wb.save(path)
    return path
